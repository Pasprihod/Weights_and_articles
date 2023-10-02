import requests
import pandas as pd
import os
from PyQt5.QtCore import QTimer
from PyQt5.QtWidgets import QMessageBox
from ultralytics import YOLO
import pickletools
import json
import datetime
from pyzbar import pyzbar # штрих-код
from openpyxl import load_workbook
from openpyxl.styles import (
                        PatternFill,
                        Alignment,
                        Font,
                        Border,
                        Side
                        )
from openpyxl.formatting.rule import Rule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from functions_service import *

CLASSES_ALL = ["табло весов", "наименование", "транс", "групп", "изделие", "штрихкод", "артикул", "дата",
                           "мануал", "сделано в", "арт_колво"]
CLASSES_NUMBERS = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", ".", "пустота"]
PATH_TO_SAVE_NUMBERS = r'temp\numbers'
PATH_TO_SAVE_TEXT = r'temp\images'
PATH_TO_DATABASE = r'data_base\files processed'
MODEL_ALL_PATH = r'models\all_model.pt'
MODEL_NUMBER_PATH = r'models\numbers_model.pt'
model_all = YOLO(MODEL_ALL_PATH)
model_number = YOLO(MODEL_NUMBER_PATH)

# распознавание текста ocr.space
def text_recog(filename, overlay=False, api_key='K86770842088957', language='eng'):
    """ OCR.space API request with local file.
        Python3.5 - not tested on 2.7
    :param filename: Your file path & name.
    :param overlay: Is OCR.space overlay required in your response.
                    Defaults to False.
    :param api_key: OCR.space API key.
                    Defaults to 'helloworld'.
    :param language: Language code to be used in OCR.
                    List of available language codes can be found on https://ocr.space/OCRAPI
                    Defaults to 'en'.
    :return: Result in JSON format.
    """

    payload = {'isOverlayRequired': overlay,
               'apikey': api_key,
               'language': language,
               'OCREngine' : 2}
    with open(filename, 'rb') as f:
        r = requests.post('https://api.ocr.space/parse/image',
                          files={filename: f},
                          data=payload,
                          )
    return  r.content.decode()


# распознавание текста на отдельной картинке
def recog_1_field(image):
    result = ''
    path_to_field_image = os.path.join(PATH_TO_SAVE_TEXT,'1.jpg')
    cv2.imwrite(path_to_field_image, image)
    string = text_recog(filename=path_to_field_image)
    data = json.loads(string) # string to json
    for i in data['ParsedResults'][0]['TextOverlay']['Lines']:
        for j in i['Words']:
            result += j['WordText'] + ' '
    print(result)
    os.remove(path_to_field_image)
    return result

# распознавание массы на табло
def weights_recog(image_cropped, file):
    (h, w) = image_cropped.shape[:2]
    if h > w:
        image_cropped = cv2.rotate(image_cropped, cv2.ROTATE_90_CLOCKWISE)
    image_cropped_path = PATH_TO_SAVE_NUMBERS + '/'+ file[:-4] + '_tablo.jpg'
    # image_cropped_path = PATH_TO_SAVE_NUMBERS + '/' + '_tablo.jpg'
    cv2_imencode(image_cropped_path, image_cropped)
    # cv2.imwrite(image_cropped_path, image_cropped)
    # cv2.imencode(".jpg", image_cropped)[1].tofile(image_cropped_path)
    result_numbers = model_number(image_cropped_path, conf = 0.5)
    boxes_numbers = result_numbers[0].boxes
    bbox_coordinates_number = boxes_numbers.xyxy.tolist()
    number = boxes_numbers.cls.tolist()
    number_conf = boxes_numbers.conf.tolist()
    conf_dot_max = 0
    for i_number, conf in enumerate(number_conf):
        if int(number[i_number]) == 10 and conf > conf_dot_max:
            conf_dot_max = conf
            maxconf_dot_number = i_number

    # print(bbox_coordinates_number)
    number_xmin = []
    number_final = ''
    for i_number, bbox_number in enumerate(bbox_coordinates_number):
        x_min, y_min, _, _ = bbox_number
        if int(number[i_number]) == 10 and number_conf[i_number] == conf_dot_max or int(number[i_number]) != 10:
            number_xmin.append((str(int(number[i_number])), x_min))
    number_xmin = sorted(number_xmin, key=lambda x: x[1])
    # print(number_xmin)
    if number_xmin[-1][0] == '11':
        number_xmin.reverse()
        for element in number_xmin:
            if int(element[0]) != 11:
                number_final += CLASSES_NUMBERS[int(element[0])]
        number_final = change_elements(number_final, '6', '9')
        number_final = change_elements(number_final, '9', '6')
    else:
        for element in number_xmin:
            if int(element[0]) != 11:
                number_final += CLASSES_NUMBERS[int(element[0])]
    # image_cropped = cv2.cvtColor(image_cropped, cv2.COLOR_BGR2RGB)
    if len(number_final) == 4 and '.' not in number_final:
        number_final = number_final[:2] + '.' + number_final[2:]
    if len(number_final) == 3 and '.' not in number_final:
        number_final = number_final[0] + '.' + number_final[1:]
    # image_cropped = cv2.cvtColor(image_cropped, cv2.COLOR_BGR2RGB)
    print(number_final)
    return number_final


# обнаружение полей
def make_items_images(BATCH_PATH):
    items_images = {} # ключ - путь к файлу, значение - словарь из параметров CLASSES_ALL
    files_path = [] # список путей к файлам с изображениями в одной папке или поартикулярно и новое название фотки в базе
    for item in sorted(os.listdir(BATCH_PATH)):
        item_path = os.path.join(BATCH_PATH, item)
        if os.path.isdir(item_path):
            for photo in sorted(os.listdir(item_path)):
                item_photo = item + '_' + photo
                file_path = os.path.join(item_path,photo) # путь к фотке
                files_path.append((file_path, item_photo))
        else:
            item_photo = BATCH_PATH.split(os.sep)[-1] + '_' + item
            files_path.append((item_path, item_photo))

    for file_path, item_photo in files_path:
                # item_photo = item + '_' + photo
        try:
        # if photo.lower().endswith('.jpg') or photo.lower().endswith('.png') or photo.lower().endswith('.tif'):
            print(file_path, os.path.isfile(file_path))
            items_images[item_photo] = {"вес": None,
                                        "наименование":'',
                                        "транс":'',
                                        "групп":'',
                                        "изделие":'',
                                        "штрихкод":'',
                                        "артикул":'',
                                        "дата":'',
                                        "мануал": None,
                                        "сделано в":'',
                                        "арт_колво": ''}
            # # открываем изображение
            # img = cv2.imdecode(np.fromfile(file_path, dtype=np.uint8), cv2.IMREAD_COLOR)
            # img = cv2.cvtColor(img,cv2.COLOR_BGR2RGB)
            # plt.imshow(img)
            # plt.show()
            # # поворачиваем изображение, чтобы большая часть текста располагалась горизонтально
            # result_obj = model_obj(file_path, conf = 0.5)
            # # print('Дошли сюда1', result_obj)
            # if result_obj:
            #     boxes_obj = result_obj[0].boxes
            #     bbox_coordinates_obj = boxes_obj.xyxy.tolist()
            #     xmin1, ymin1, xmax1, ymax1 = bbox_coordinates_obj[0]
            #     img_cropped = img[int(ymin1):int(ymax1), int(xmin1):int(xmax1),  :]
            #     angle = horiz(img_cropped)
            #     rows, cols = img.shape[:2]
            #     M = cv2.getRotationMatrix2D((cols // 2, rows // 2), angle, 1)
            #     img = cv2.warpAffine(img, M, (cols, rows))
            #     success, im_buf_arr = cv2.imencode('.jpg', img)
            #     if success:
            #         im_buf_arr.tofile(file_path)


            # распознаем поля
            img = cv2.imdecode(np.fromfile(file_path, dtype=np.uint8), cv2.IMREAD_COLOR)
            img = cv2.cvtColor(img,cv2.COLOR_BGR2RGB)
            results_model_all = model_all(file_path, conf=0.5)   # распознаем все поля на снимке

            boxes = results_model_all[0].boxes                      # выделяем результат: bbox, класс
            item_class = boxes.cls.tolist()                             # выделяем  класс в список
            bbox_coordinates = boxes.xyxy.tolist()                # выделяем  bbox в список


            for i,bbox in enumerate(bbox_coordinates):
                xmin, ymin, xmax, ymax = bbox
                image_cropped = img[int(ymin):int(ymax), int(xmin):int(xmax),  :]
                (h, w) = image_cropped.shape[:2]
                if h > w:
                    image_cropped = cv2.rotate(image_cropped, cv2.ROTATE_90_CLOCKWISE)
                angle = 0
                if CLASSES_ALL[int(item_class[i])] == 'табло весов':

                    # # поворачиваем табло
                    # angle = horiz(image_cropped)
                    # rows, cols = image_cropped.shape[:2]
                    # M = cv2.getRotationMatrix2D((cols // 2, rows // 2), angle, 1)
                    # image_cropped = cv2.warpAffine(image_cropped, M, (cols, rows))

                    number_final = weights_recog(image_cropped, photo)
                    if number_final:
                        items_images[item_photo]["вес"] = float(number_final)
                    else:
                        items_images[item_photo]["вес"] = 'Не распознано'
                elif CLASSES_ALL[int(item_class[i])] == 'наименование':
                    items_images[item_photo]["наименование"] = image_cropped
                elif CLASSES_ALL[int(item_class[i])] == 'транс':
                    items_images [item_photo]["транс"] = image_cropped
                elif CLASSES_ALL[int(item_class[i])] == 'групп':
                    items_images[item_photo]["групп"] = image_cropped
                elif CLASSES_ALL[int(item_class[i])] == 'штрихкод':
                    code = decode_pyzbar(image_cropped)
                    items_images[item_photo]["штрихкод"] = code
                elif CLASSES_ALL[int(item_class[i])] == 'артикул':
                    items_images[item_photo]["артикул"] = image_cropped
                elif CLASSES_ALL[int(item_class[i])] == 'дата':
                    items_images[item_photo]["дата"] = image_cropped
                elif CLASSES_ALL[int(item_class[i])] == 'мануал':
                    items_images[item_photo]["мануал"] = True
                elif CLASSES_ALL[int(item_class[i])] == 'сделано в':
                    items_images[item_photo]["сделано в"] = image_cropped
                elif CLASSES_ALL[int(item_class[i])] == 'изделие':
                    items_images[item_photo]["изделие"] = image_cropped
                elif CLASSES_ALL[int(item_class[i])] == 'арт_колво':
                    items_images[item_photo]["арт_колво"] = image_cropped
        except:
            pass

    return items_images


# подсчет количества листов мануалов
def get_manuals_number(items_images, unique_items):
    manuals = {item: [[],0] for item in unique_items}
    for item_photo, fields in items_images.items():
        item = item_photo.split('_')[0]
        if fields['мануал']:
            manuals[item][0].append(item_photo)
            manuals[item][1] +=1
    return manuals # словарь по item с количеством листов мануалов

# распознавание шрихкодов
def decode_pyzbar(image):
    # decodes all barcodes from an image
    decoded_objects = pyzbar.decode(image)
    data = []
    for obj in decoded_objects:
        # draw the barcode
        # print(f"Обнаружен штрих-код:\n{obj}")
        # image = draw_barcode(obj, image)
        # print barcode type & data
        # print("Тип:", obj.type)
        print("Штрихкод:", obj.data.decode("utf-8"))
        print()
        data.append(obj.data.decode("utf-8"))

    return data[0] if data else ''


# создание словарей со списками по артикулам для записи в эксель
def get_trans_group_product_manuals(items_images, unique_items):
    trans = {item: [[], '','','','','',''] for item in unique_items}
    group = {item: [[], '','','','','',''] for item in unique_items}
    product = {item: [[], '','','','','',''] for item in unique_items}

    n = 1 # количество запросов на распознавание текста
    for item_photo, fields in items_images.items():
        item = check_item(unique_items, item_photo) # ищем артикул из экселя, который присутствует в названии файла или папки
        art = ''
        if not item: # если в названии файла или папки не нашелся артикул из экселя, распознаем артикул на изо и сверяем
            if len(fields['артикул']):
                    file_field_name = f'{item_photo[:-4]}_артикул.txt'
                    field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                    text, n = check_and_get_field_text(fields['артикул'], field_file_path, n)
                    art = correct_article(text)
                    if not art and len(fields['арт_колво']):
                        field_file_path = os.path.join(PATH_TO_DATABASE, f'{item_photo}_арт_колво.txt')
                        text, n = check_and_get_field_text(fields['арт_колво'], field_file_path, n)
                        art = correct_article(text.split('/')[0])
                    # print(file_field_name, art)
                    if art in unique_items:
                        item = art
        if item: # если найденный артикул совпадает с каким-то из экселя
            # ТРАНС
            if len(fields['транс']) or len(fields['арт_колво']):
                # создаем список из фото с транспортной упаковкой
                trans[item][0].append(item_photo)
                # артикул
                if art:
                    trans[item][1] = art
                else:
                    if len(fields['артикул']):
                        if not trans[item][1] or trans[item][1] == 'Не распознано':
                            art = 'Не распознано'
                            file_field_name = f'{item_photo[:-4]}_артикул.txt'
                            field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                            text, n = check_and_get_field_text(fields['артикул'], field_file_path, n)
                            art = correct_article(text)
                            if not art and len(fields['арт_колво']):
                                field_file_path = os.path.join(PATH_TO_DATABASE, f'{item_photo}_арт_колво.txt')
                                text, n = check_and_get_field_text(fields['арт_колво'], field_file_path, n)
                                art = correct_article(text.split('/')[0])
                            # print(file_field_name, art)
                            trans[item][1] = art

                 # наименование
                if len(fields['наименование']):
                    if not trans[item][2] or trans[item][2] == 'Не распознано':
                        name = 'Не распознано'
                        file_field_name = f'{item_photo[:-4]}_наименование.txt'
                        field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                        name, n = check_and_get_field_text(fields['наименование'], field_file_path, n)
                        trans[item][2] = name

                # кол-во
                if not trans[item][3] or trans[item][3] == 'Не распознано':
                    file_field_name = f'{item_photo[:-4]}_транс.txt'
                    field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                    text, n = check_and_get_field_text(fields['транс'], field_file_path, n)
                    trans_count_text = extract_text_number(text)
                    if not trans_count_text and len(fields['арт_колво']):
                        field_file_path = os.path.join(PATH_TO_DATABASE, f'{item_photo}_арт_колво.txt')
                        text, n = check_and_get_field_text(fields['арт_колво'], field_file_path)
                        trans_count_text = extract_text_number(text.split('/')[1])
                    if trans_count_text:
                        trans_count = int(trans_count_text)
                    else:
                        trans_count = 'Не распознано'
                    trans[item][3] = trans_count
                    # print(file_field_name, trans_count)

                # сделано в
                if len(fields['сделано в']):
                    if not trans[item][4] or trans[item][4] == 'Не распознано':
                        made = 'Не распознано'
                        file_field_name = f'{item_photo[:-4]}_сделано в.txt'
                        field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                        made, n = check_and_get_field_text(fields['сделано в'], field_file_path, n)
                        trans[item][4] = made

                # штрихкод
                if fields['штрихкод']:
                    if not trans[item][5] or trans[item][5] == 'Не распознано':
                        trans[item][5] = fields['штрихкод']
                if fields['вес']:
                    if not trans[item][6] or trans[item][6] == 'Не распознано':
                        trans[item][6] = fields['вес']

             # ГРУПП
            if len(fields['групп']) and not fields['штрихкод'].startswith('46'):
                # создаем список из фото с транспортной упаковкой
                group[item][0].append(item_photo)
                # артикул
                if art:
                    if not group[item][1] or group[item][1] == 'Не распознано':
                        group[item][1] = art
                else:
                    if len(fields['артикул']):
                        if not group[item][1] or group[item][1] == 'Не распознано':
                            art = 'Не распознано'
                            file_field_name = f'{item_photo[:-4]}_артикул.txt'
                            field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                            text, n = check_and_get_field_text(fields['артикул'], field_file_path, n)
                            art = correct_article(text)
                            group[item][1] = art

                 # наименование
                if len(fields['наименование']):
                    if not group[item][2] or group[item][2] == 'Не распознано':
                        name = 'Не распознано'
                        file_field_name = f'{item_photo[:-4]}_наименование.txt'
                        field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                        name, n = check_and_get_field_text(fields['наименование'], field_file_path, n)
                        group[item][2] = name

                # кол-во
                if not group[item][3] or group[item][3] == 'Не распознано':
                    file_field_name = f'{item_photo[:-4]}_групп.txt'
                    field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                    text, n = check_and_get_field_text(fields['групп'], field_file_path, n)
                    group_count_text = extract_text_number(text)
                    if group_count_text:
                        group_count = int(group_count_text)
                    else:
                        group_count = 'Не распознано'
                    group[item][3] = group_count


                # сделано в
                if len(fields['сделано в']):
                    if not group[item][4] or group[item][4] == 'Не распознано':
                        made = 'Не распознано'
                        file_field_name = f'{item_photo[:-4]}_сделано в.txt'
                        field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                        made, n = check_and_get_field_text(fields['сделано в'], field_file_path, n)
                        group[item][4] = made

                # дата
                if len(fields['дата']):
                    if not group[item][5] or group[item][5] == 'Не распознано':
                        date = 'Не распознано'
                        file_field_name = f'{item_photo[:-4]}_дата.txt'
                        field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                        date, n = check_and_get_field_text(fields['дата'], field_file_path, n)
                        group[item][5] = date

                # штрихкод
                if fields['штрихкод']:
                    if not group[item][6] or group[item][6] == 'Не распознано':
                        group[item][6] = fields['штрихкод']

             # ИЗДЕЛИЕ

            if len(fields['изделие']) or fields['штрихкод'].startswith('46'):
                # создаем список из фото с транспортной упаковкой
                product[item][0].append(item_photo)
                # артикул
                if art:
                    if not product[item][1] or product[item][1] == 'Не распознано':
                        product[item][1] = art
                else:
                    if len(fields['артикул']):
                        if not product[item][1] or product[item][1] == 'Не распознано':
                            art = 'Не распознано'
                            file_field_name = f'{item_photo[:-4]}_артикул.txt'
                            field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                            text, n = check_and_get_field_text(fields['артикул'], field_file_path, n)
                            art = correct_article(text)
                            product[item][1] = art

                 # наименование
                if len(fields['наименование']):
                    if not product[item][2] or product[item][2] == 'Не распознано':
                        name = 'Не распознано'
                        file_field_name = f'{item_photo[:-4]}_наименование.txt'
                        field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                        name, n = check_and_get_field_text(fields['наименование'], field_file_path, n)
                        product[item][2] = name

                # сделано в
                if len(fields['сделано в']):
                    if not product[item][3] or product[item][3] == 'Не распознано':
                        made = 'Не распознано'
                        file_field_name = f'{item_photo[:-4]}_сделано в.txt'
                        field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                        made, n = check_and_get_field_text(fields['сделано в'], field_file_path, n)
                        product[item][3] = made

                # дата
                if len(fields['дата']):
                    if not product[item][4] or product[item][4] == 'Не распознано':
                        date = 'Не распознано'
                        file_field_name = f'{item_photo[:-4]}_дата.txt'
                        field_file_path = os.path.join(PATH_TO_DATABASE, file_field_name)
                        date, n = check_and_get_field_text(fields['дата'], field_file_path, n)
                        product[item][4] = date

                # штрихкод
                if fields['штрихкод']:
                    if not product[item][5] or product[item][5] == 'Не распознано':
                        product[item][5] = fields['штрихкод']
                # вес
                if fields['вес']:
                    if not product[item][6] or product[item][6] == 'Не распознано':
                        product[item][6] = fields['штрихкод']

    manuals = get_manuals_number(items_images, unique_items)
    return trans, group, product,manuals, n

# список артикулов из Экселя
def get_unique_items(PATH_TO_EXCEL_ORIGIN):
    excel = pd.read_excel(PATH_TO_EXCEL_ORIGIN)
    row,column = np.where(excel == 'ARTICLE')
    art_column = excel.columns[column[0]]
    excel = excel.fillna(' ')
    id = excel.index[excel[art_column].str.contains('-')]
    return excel.loc[id,art_column].tolist()

# проверка, распознавалось ли ранее данное поле, если нет, то распознавание
def check_and_get_field_text(field, field_file_path, n, PATH=PATH_TO_DATABASE): #field - картинка
    file_name = field_file_path.split(os.sep)[-1]
    if file_name not in os.listdir(PATH):
        if not n % 180:
            message_box = QMessageBox()
            message_box.setWindowTitle('Предупреждение')
            current_time = datetime.datetime.now()
            current_time = current_time.strftime("%H:%M")
            message_box.setText(
                f'Достигнут предел в 180 запросов на распознавание.\nТекущее время: {current_time}.\nПодождите 1 час...')
            message_box.setStandardButtons(QMessageBox.NoButton)
            QTimer.singleShot(3650000, message_box.accept) # 3650 сек
            message_box.exec_()
        text = recog_1_field(field)
        n += 1
        with open(field_file_path, 'w', encoding='utf-8') as f:
            f.write(text)
    else:
        with open(field_file_path, 'r', encoding='utf-8') as f:
            text = f.read()
    return text, n

# ЭКСЕЛЬ

# переименование содержания ячеек с Nan в строке с названием колонок
def rename_nan_columns(df):
    col_new = []
    col = list(df.columns)
    for i in range(len(col)):
        if col[i] is np.nan:
            col_new.append(str(i))
        else:
            col_new.append(col[i])
    df.columns = col_new
    return df


# поиск координат поля на активном листе экселя
def find_cell_in_excel(ws, text):
    row_max = ws.max_row  # Получаем количество столбцов
    #print(type(row_max))
    column_max = ws.max_column  # Получаем количество строк

    # print('В файле:', path_to_file, '\n Cтолбцов:', row_max, '\n Колонок:', column_max)

    row_min = 1 #Переменная, отвечающая за номер строки
    column_min = 1 #Переменная, отвечающая за номер столбца

    while column_min <= column_max:
        row_min_min = row_min
        row_max_max = row_max
        while row_min_min <= row_max_max:
            row_min_min = str(row_min_min)

            word_column = get_column_letter(column_min)
            word_column = str(word_column)
            word_cell = word_column + row_min_min

            data_from_cell = ws[word_cell].value
            data_from_cell = str(data_from_cell)
            #print(data_from_cell)

            if text in data_from_cell:
                # print(word_column + row_min_min)
                return word_column, row_min_min
            row_min_min = int(row_min_min)
            row_min_min = row_min_min + 1
        column_min = column_min + 1

# форматирование и запись в эксель
def to_excel(PATH_TO_EXCEL_ORIGIN, PATH_TO_EXCEL_RESULT, unique_items, trans, group, product, manuals):
    try:
        if PATH_TO_EXCEL_ORIGIN.endswith('.xlsx'):
            excel = pd.read_excel(PATH_TO_EXCEL_ORIGIN)
            idx_columns = np.where(excel == 'ARTICLE')[0][0]
            excel.columns = excel.loc[idx_columns]
            excel = rename_nan_columns(excel)

            # ключевые поля в экселе
            trans_weight_excel = 'GR. WT\n(KG)'
            art_excel = 'ARTICLE'
            count_trans = 'UNITS/\nCTN'
            count_grup = 'Sale\nBag'
            count_grup_1 = 'Small\nBag'
            last_column = 'depth'

            new_columns = ['Транс. Артикул',
                           'Транс. Наименование',
                           'Транс. Количество',
                           'Транс. Производитель',
                           'Транс. Штрихкод',
                           'Транс. Вес',
                           'Групп. Артикул',
                           'Групп. Наименование',
                           'Групп. Количество',
                           'Групп. Производитель',
                           'Групп. Дата',
                           'Групп. Штрихкод',
                           'Изделие Артикул',
                           'Изделие Наименование',
                           'Изделие Производитель',
                           'Изделие Дата',
                           'Изделие Штрихкод',
                           'Изделие Вес',
                           'Кол-во стр. мануалов']

            excel[new_columns] = np.nan
            excel[new_columns] = excel[new_columns].astype('object')
            excel.loc[idx_columns, new_columns] = new_columns
            for item in unique_items:
                excel_tmp = excel.copy().fillna(' ')
                id = excel.index[excel_tmp[art_excel].str.contains(item)]  # (excel[art_excel] == trans[k][1]) |
                excel.loc[id, new_columns[:6]] = trans[item][1:7]
                excel.loc[id, new_columns[6:12]] = group[item][1:7]
                excel.loc[id, new_columns[12:18]] = product[item][1:7]
                excel.loc[id, new_columns[18]] = manuals[item][1]
            font = Font(
                name='Arial',
                size=9,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000'
            )

            alignment = Alignment(
                horizontal='center',
                vertical='center',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0
            )

            wb = load_workbook(PATH_TO_EXCEL_ORIGIN)
            ws = wb.active

            # поиск букв столбцов ключевых полей
            trans_weight_letter, _ = find_cell_in_excel(ws, trans_weight_excel)
            art_letter, _ = find_cell_in_excel(ws, art_excel)
            count_trans_letter, _ = find_cell_in_excel(ws, count_trans)
            count_grup_letter, _ = find_cell_in_excel(ws, count_grup)
            count_grup_1_letter, _ = find_cell_in_excel(ws, count_grup_1)
            last_column_letter, row_last = find_cell_in_excel(ws, last_column)

            column_number_last = ws[f'{last_column_letter}{row_last}'].column
            r_start = idx_columns + 2
            c = column_number_last + 1
            columns_letters = []
            r = r_start

            len_unique_items = len(unique_items)
            for col in new_columns:
                for row in range(len_unique_items + 2):
                    cell = ws.cell(row=r,
                                   column=c)  # Создаем ячейку с координатами, где показываем результаты из пандас таблицы
                    if r in (r_start, r_start + 1):
                        cell.border = Border(top=Side(border_style="medium", color="FF000000"),
                                             left=Side(border_style="medium", color="FF000000"),
                                             right=Side(border_style="medium", color="FF000000"),
                                             bottom=Side(border_style="medium", color="FF000000"))
                    elif col == new_columns[0] and r not in (r_start, r_start + 1):
                        cell.border = Border(right=Side(border_style="thin", color="FF000000"),
                                             bottom=Side(border_style="thin", color="FF000000"))
                    else:
                        cell.border = Border(top=Side(border_style="thin", color="FF000000"),
                                             left=Side(border_style="thin", color="FF000000"),
                                             right=Side(border_style="thin", color="FF000000"),
                                             bottom=Side(border_style="thin", color="FF000000"))

                    cell.value = excel.loc[idx_columns + row, col]
                    cell.font = font
                    cell.alignment = alignment
                    if cell.column_letter not in columns_letters:
                        columns_letters.append(cell.column_letter)
                    r += 1
                c += 1
                r = r_start

            # определение начальной и конечной буквы столбцов
            column_start = columns_letters[0]
            column_finish = columns_letters[-1]

            # Подбор ширины столбцов по содержимому
            # словарь с размерами столбцов
            cols_dict = {}
            # проходимся по всем строкам документа
            for row in ws.rows:
                # теперь по ячейкам каждой строки
                for i in range(column_number_last, ws.max_column):
                    cell = row[i]
                    # получаем букву текущего столбца
                    letter = cell.column_letter
                    # если в ячейке записаны данные
                    if cell.value:

                        # вычисляем количество символов, записанных в ячейку
                        len_cell = len(str(cell.value))
                        # длинна колонки по умолчанию, если буква
                        # текущего столбца отсутствует в словаре `cols_dict`
                        len_cell_dict = 0
                        # смотрим в словарь c длинами столбцов
                        if letter in cols_dict:
                            # если в словаре есть буква текущего столбца
                            # то извлекаем соответствующую длину
                            len_cell_dict = cols_dict[letter]

                        # если текущая длина данных в ячейке
                        #  больше чем длинна из словаря
                        if len_cell > len_cell_dict:
                            # записываем новое значение ширины этого столбца
                            cols_dict[letter] = len_cell
                            ###!!! ПРОБЛЕМА АВТОМАТИЧЕСКОЙ ПОДГОНКИ !!!###
                            ###!!! расчет новой ширины колонки (здесь надо подгонять) !!!###
                            new_width_col = len_cell * 1.3
                            # применение новой ширины столбца
                            ws.column_dimensions[cell.column_letter].width = new_width_col

            # Назначение ширины стобцов для наименования транс, групп, изделия
            ws.column_dimensions[columns_letters[1]].width = 40
            ws.column_dimensions[columns_letters[7]].width = 40
            ws.column_dimensions[columns_letters[13]].width = 40

            green = "5cb800"
            red = "9C0006"
            yellow = 'ffff00'
            red_text = Font(color=red)
            greenFill = PatternFill('solid', bgColor=green)
            redFill = PatternFill('solid', bgColor=red)
            yellowFill = PatternFill('solid', bgColor=yellow)
            operator_fill = ('=', greenFill), ("!=", redFill)
            last_row = r_start + len_unique_items + 1

            diff_style_green = DifferentialStyle(fill=greenFill)
            diff_style_red = DifferentialStyle(fill=redFill)

            # транс артикул, транс количество,  групп артикул, изделие артикул
            number_fields = (0, art_letter), (2, count_trans_letter), (6, art_letter), (12, art_letter)
            for number, field in number_fields:
                formula_equal = f'AND({columns_letters[number]}{r_start + 2} <> "", {columns_letters[number]}{r_start + 2} = {field}{r_start + 2})'
                formula_unequal = f'AND({columns_letters[number]}{r_start + 2} <> "", {columns_letters[number]}{r_start + 2} <> {field}{r_start + 2})'
                rule_equal = Rule(type="expression", formula=[formula_equal], dxf=diff_style_green)
                rule_unequal = Rule(type="expression", formula=[formula_unequal], dxf=diff_style_red)
                ws.conditional_formatting.add(
                    f'{columns_letters[number]}{r_start + 2}:{columns_letters[number]}{last_row}', rule_equal)
                ws.conditional_formatting.add(
                    f'{columns_letters[number]}{r_start + 2}:{columns_letters[number]}{last_row}', rule_unequal)

            # групп количество сравненние с двумя колонками
            diff_style_green = DifferentialStyle(fill=greenFill)
            diff_style_red = DifferentialStyle(fill=redFill)
            formula_equal = f'AND({columns_letters[8]}{r_start + 2} <> "", OR({columns_letters[8]}{r_start + 2}={count_grup_letter}{r_start + 2},{columns_letters[8]}{r_start + 2}={count_grup_1_letter}{r_start + 2}))'
            formula_unequal = f'AND({columns_letters[8]}{r_start + 2} <> "", AND({columns_letters[8]}{r_start + 2}<>{count_grup_letter}{r_start + 2},{columns_letters[8]}{r_start + 2}<>{count_grup_1_letter}{r_start + 2}))'
            rule_equal = Rule(type="expression", formula=[formula_equal], dxf=diff_style_green)
            rule_unequal = Rule(type="expression", formula=[formula_unequal], dxf=diff_style_red)
            ws.conditional_formatting.add(f'{columns_letters[8]}{r_start + 2}:{columns_letters[8]}{last_row}',
                                          rule_equal)
            ws.conditional_formatting.add(f'{columns_letters[8]}{r_start + 2}:{columns_letters[8]}{last_row}',
                                          rule_unequal)

            dxf1 = DifferentialStyle(fill=yellowFill)
            for c in columns_letters:
                rule2 = Rule(type="containsText", operator="containsText", text="Не распознано", dxf=dxf1)
                rule2.formula = [f'NOT(ISERROR(SEARCH("Не распознано",{c}{r_start + 2})))']
                ws.conditional_formatting.add(f'{c}{r_start + 2}:{c}{last_row}', rule2)
                # ws.conditional_formatting.add(f'{c}{r_start+2}:{c}{last_row}',
                #     FormulaRule(formula=[f'{c}{r_start+2}=Не распознано'],  fill=yellowFill))

                ws.conditional_formatting.add(f'{c}{r_start + 2}:{c}{last_row}',
                                              CellIsRule(operator='=', formula=['Не распознано'], fill=yellowFill))

            # форматирование весов транс (зеленым разница < 5 %)

            # # составляем правило
            rule_green = Rule(type="expression", dxf=diff_style_green)
            # логика правила, которая соответствует `type="expression"`
            rule_green.formula = [
                f"ABS({trans_weight_letter}{r_start + 2}-{columns_letters[5]}{r_start + 2})/{trans_weight_letter}{r_start + 2}*100 < 5"]
            # добавляем правило форматирования
            ws.conditional_formatting.add(f"{columns_letters[5]}{r_start + 2}:{columns_letters[5]}{last_row}",
                                          rule_green)

            # # составляем правило
            rule_red = Rule(type="expression", dxf=diff_style_red)
            # логика правила, которая соответствует `type="expression"`
            rule_red.formula = [
                f"ABS({trans_weight_letter}{r_start + 2}-{columns_letters[5]}{r_start + 2})/{trans_weight_letter}{r_start + 2}*100 >= 5"]
            # добавляем правило форматирования
            ws.conditional_formatting.add(f"{columns_letters[5]}{r_start + 2}:{columns_letters[5]}{last_row}",
                                          rule_red)

            for c in range(column_number_last + 1, ws.max_column + 1):
                ws.merge_cells(start_row=r_start, start_column=c, end_row=r_start + 1, end_column=c)
            new_excel_path = os.path.join(PATH_TO_EXCEL_RESULT, f'{PATH_TO_EXCEL_ORIGIN.split(os.sep)[-1][:-5]}_REPORT.xlsx')
            wb.save(new_excel_path)
            file_name = new_excel_path.split('/')[-1]
            return f'Создан файл {file_name} в папке {PATH_TO_EXCEL_RESULT}'
        else:
            f'Загружаемый файл {PATH_TO_EXCEL_RESULT.split(os.sep)[-1]} имеет недопустимый формат .xls. Пересохраните его в формате .xlsx'
    except FileNotFoundError:
        return f'Файл {PATH_TO_EXCEL_ORIGIN} не найден!'
    except OSError:
        return f'Директория {PATH_TO_EXCEL_RESULT} не найдена!'